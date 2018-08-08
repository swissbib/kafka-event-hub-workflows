from openpyxl import load_workbook
from simple_elastic import ElasticIndex

# TODO: Rework how this works and think about how I need the data?

wb = load_workbook('data/VERZ_DSV01-Ausleihen-Vormerkungen_20180802_bmt.xlsx')
ws = wb['vor-1900-vormerkungen']

index = ElasticIndex('reservations', 'record')

for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
    doc = dict()
    system_number = str(row[0].value)
    while len(system_number) != 9:
        system_number = '0' + system_number
    doc['system_number'] = system_number
    doc['reservations'] = row[1].value
    doc['year'] = str(row[2].value)
    index.index_into(doc, doc['system_number'] + doc['year'])

ws = wb['vor-1900-ausleihen']

for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
    doc = dict()
    system_number = str(row[0].value)
    while len(system_number) != 9:
        system_number = '0' + system_number
    doc['system_number'] = system_number
    doc['loans'] = row[1].value
    doc['year'] = str(row[2].value)
    record = index.get(doc['system_number'] + doc['year'])
    if record is not None:
        doc['reservations'] = record['reservations']
        index.index_into(doc, doc['system_number'] + doc['year'])
    else:
        index.index_into(doc, doc['system_number'] + doc['year'])
