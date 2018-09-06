from openpyxl import load_workbook
from simple_elastic import ElasticIndex

if __name__ == '__main__':

    wb = load_workbook('data/VERZ_DSV01-Ausleihen-Vormerkungen_20180802_bmt.xlsx')
    ws = wb['vor-1900-vormerkungen']

    index = ElasticIndex('aleph-dsv01-data', 'hits')

    all_data = dict()

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
        doc = dict()
        system_number = str(row[0].value)
        while len(system_number) != 9:
            system_number = '0' + system_number

        if system_number not in all_data:
            all_data[system_number] = dict()
        if 'reservations' not in all_data[system_number]:
            all_data[system_number]['reservations'] = dict()
        all_data[system_number]['reservations'][str(row[2].value)] = row[1].value

    ws = wb['vor-1900-ausleihen']

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
        doc = dict()
        system_number = str(row[0].value)
        while len(system_number) != 9:
            system_number = '0' + system_number
        if system_number not in all_data:
            all_data[system_number] = dict()
        if 'loans' not in all_data[system_number]:
            all_data[system_number]['loans'] = dict()
        all_data[system_number]['loans'][str(row[2].value)] = row[1].value

    list_of_data = list()
    for system_number in all_data:
        data = all_data[system_number]
        for t in ['loans', 'reservations']:
            if t in data:
                total = 0
                for y in ['2016', '2017', '2018']:
                    if y not in data[t]:
                        data[t][y] = 0
                    else:
                        total += data[t][y]
                data[t]['total'] = total
            else:
                data[t] = {'2016': 0, '2017': 0, '2018': 0, 'total': 0}

        data['identifier'] = system_number
        list_of_data.append(data)

    index.bulk(list_of_data, 'identifier')
